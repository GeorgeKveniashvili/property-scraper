import openpyxl
import time
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Default location of the excel file to write to
excel_file_location = 'Properties.xlsx'

# Class which is responsible for scraping the zoopla.co.uk
# Method which does the scraping is called do_scrape
class ZooplaScraper:
    # Editable parameters for filtering the properties
    location = 'London'
    min_price = '80000'
    max_price = '90000'
    
    # Static parameters that should not be changed
    website_url_getter = 'https://www.zoopla.co.uk/api/search/resolver/?price_frequency=per_month&results_sort=newest_listings&new_homes=include&retirement_homes=true&shared_ownership=true&include_shared_accommodation=true&search_source=for-sale&section=for-sale&view_type=list&price_max={prop_max_price}&price_min={prop_min_price}&q={prop_query1}&orig_q={prop_query2}'.format(prop_max_price=max_price, prop_min_price=min_price, prop_query1=location, prop_query2=location)
    # Starting page number of the website
    page_number = 1
    # Name of the Excel sheet where the data is inserted
    excel_sheet = 'zoopla'
    # Row index where insterting should start
    row_index = 2

    # Method which opens the website and scrapes the properties
    def do_scrape(self):
        print('Scraping zoopla.co.uk')
        
        # Open Chrome and load the webpage which returns the actual property search URL
        dr = webdriver.Chrome()
        dr.maximize_window()
        dr.get(self.website_url_getter)

        # Wait 60 seconds until the response is returned
        # Used for allowing solving the reCAPTCHA
        try:
            WebDriverWait(dr, 60).until(
                EC.presence_of_element_located((By.TAG_NAME, 'pre'))
                )
        except:
            print('Could not load zoopla')
            return

        # Get the property search URL
        website_uri = dr.find_element(By.TAG_NAME, 'pre').text

        # Loop over every webpage
        while True:
            # Format the URL and load the page
            website_url = 'https://www.zoopla.co.uk{uri}&pn={page_number}'.format(uri=website_uri, page_number=self.page_number)
            dr.get(website_url)

            # Wait 60 seconds until the response is returned
            # Used for allowing solving the reCAPTCHA
            try:
                WebDriverWait(dr, 60).until(
                    EC.presence_of_element_located((By.CLASS_NAME, '_1maljyt1'))
                    )
            except:
                print('Could not load zoopla')
                break

            # Get HTML of the webpage
            page_soup = BeautifulSoup(dr.page_source, 'html.parser')
            # Get list of URLs of the individual properties
            page_listings = page_soup.find_all('a', class_='_1maljyt1')

            # Iterate over every property URL
            for listing in page_listings:
                # Get text of the URL
                listing_url = self.convert_url(listing['href'])
                # Get full description of the property listing
                listing_desc = listing.find('div', class_='_1ankud50')

                # Try parsing the address from the description
                try:
                    listing_address = listing_desc.find('h3').getText()
                except:
                    listing_address = 'Could not get address'
                # Try parsing the title from the description
                try:
                    listing_title = listing_desc.find('h2').getText()
                except:
                    listing_title = 'Could not get title'
                # Try parsing the price from the description
                try:
                    listing_price = listing.find('p', class_='_170k6632').getText()
                except:
                    listing_price = 'Could not get price'
                
                # Insert the data into the Excel file
                write_excel(self.excel_sheet, self.row_index, listing_url, listing_address, listing_title, listing_price)
                # Increase the Excel row index for the next insertion
                self.row_index += 1
            
            # Check if the next page button is active
            # If it is, it means that there are more pages to load
            try:
                nav_menu = dr.find_element(By.CLASS_NAME, '_13wnc6k0')
                next_button = nav_menu.find_element(By.CLASS_NAME, '_1ljm00us').find_element(By.TAG_NAME, 'a')
            except:
                # Exit the method if next button can not be found
                break
            if next_button.get_attribute('aria-disabled') == 'true':
                # Exit the method if next button is disabled
                break
            
            # Increase website page index to load the next page
            self.page_number += 1
            # Wait for 1 second. Additional protection from bot detection
            time.sleep(1)
        
        # Close the browser
        dr.close()
    
    # Format the property listing URL 
    def convert_url(self, href):
        base_url = 'https://www.zoopla.co.uk'
        return base_url + href.split('/?')[0]

# Class which is responsible for scraping the rightmove.co.uk
# Method which does the scraping is called do_scrape
class RightMoveScraper:
    # Editable parameters for filtering the properties
    # The region code must be detected manually by going to www.rightmove.co.uk
    #  searching with that region and getting the region code from the URL
    # e.g. search URL https://www.rightmove.co.uk/property-for-sale/find.html?searchType=SALE&locationIdentifier=REGION%5E1498
    #  location code is written in locationIdentifier parameter and is - REGION%5E1498
    location = 'REGION%5E305'
    min_price = '350000'
    max_price = '350000'

    # Static parameters that should not be changed
    website_url = 'https://www.rightmove.co.uk/property-for-sale/find.html?locationIdentifier={prop_location}&radius=0.0&maxPrice={prop_max_price}&minPrice={prop_min_price}&includeSSTC=false'.format(prop_location=location, prop_max_price=max_price, prop_min_price=min_price)
    # Name of the Excel sheet where the data is inserted
    excel_sheet = 'rightmove'
    # Row index where insterting should start
    row_index = 2
    
    # Method which opens the website and scrapes the properties
    def do_scrape(self):
        print('Scraping rightmove.co.uk')

        # Open Chrome and load the webpage
        dr = webdriver.Chrome()
        dr.maximize_window()
        dr.get(self.website_url)
        time.sleep(1)

        # Loop over every webpage
        while True:
            # Get HTML of the webpage
            page_soup = BeautifulSoup(dr.page_source, 'html.parser')
            # Get list of URLs of the individual properties
            page_listings = page_soup.find_all('a', class_='propertyCard-anchor')

            # Iterate over every property URL
            for listing in page_listings:
                # Get text of the URL
                listing_url = self.convert_url(listing['id'])
                # Get full description of the property listing
                listing_desc = listing.find_next_sibling('div')

                # Try parsing the address from the description
                try:
                    listing_address = listing_desc.find('address', class_='propertyCard-address').getText()
                except:
                    listing_address = 'Could not get address'
                # Try parsing the title from the description
                try:
                    listing_title = listing_desc.find('span', attrs={'data-test': 'property-description'}).find('span').getText()
                except:
                    listing_title = 'Could not get title'
                # Try parsing the price from the description
                try:
                    listing_price = listing_desc.find('div', class_='propertyCard-priceValue').getText()
                except:
                    listing_price = 'Could not get price'

                # Insert the data into the Excel file
                write_excel(self.excel_sheet, self.row_index, listing_url, listing_address, listing_title, listing_price)
                # Increase the Excel row index for the next insertion
                self.row_index += 1
            
            # Check if the next page button is active
            try:
                next_button = dr.find_element(By.CLASS_NAME, 'pagination-direction--next')
            except:
                # Exit the method if next button can not be found
                break
            if not next_button.is_enabled():
                # Exit the method if next button is disabled
                break

            # Click the next button
            dr.execute_script("arguments[0].click();", next_button)
            time.sleep(1)
            # Wait until the new paig loads
            try:
                WebDriverWait(dr, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'propertyCard-anchor'))
                )
            except:
                # Exit the method if the page doesn't load in 20 seconds
                print('Could not load rightmove')
                break
        
        # Close the browser
        dr.close()

    # Format the property listing URL 
    def convert_url(self, href):
        base_url = 'https://www.rightmove.co.uk/properties/'
        return base_url + href[4:]

# Class which is responsible for scraping the gascoignehalman.co.uk
# Method which does the scraping is called do_scrape
class HalmanScraper:
    # Editable parameters for filtering the properties
    location = 'cheadle'
    min_price = '350000'
    max_price = '350000'

    # Static parameters that should not be changed
    website_url = 'https://www.gascoignehalman.co.uk/search/?showstc=on&showsold=on&instruction_type=Sale&place={prop_location}&ajax_border_miles=1&minprice={prop_min_price}&maxprice={prop_max_price}'.format(prop_location=location, prop_min_price=min_price, prop_max_price=max_price)
    # Name of the Excel sheet where the data is inserted
    excel_sheet = 'gascoignehalman'
    # Row index where insterting should start
    row_index = 2

    # Method which opens the website and scrapes the properties
    def do_scrape(self):
        print('Scraping gascoignehalman.co.uk')

        # Open Chrome and load the webpage
        dr = webdriver.Chrome()
        dr.maximize_window()
        dr.get(self.website_url)
        time.sleep(1)
        # Get the current scroll height for infinite scrolling
        previous_height = dr.execute_script('return document.body.scrollHeight')

        # Loop until the end of scrolling is reached
        while True:
            # Find element to scroll to
            scroll_to_elements = dr.find_elements(By.CLASS_NAME, 'btn-red')
            # Scroll to that element
            dr.execute_script("arguments[0].scrollIntoView();", scroll_to_elements[-1])
            # Wait for the new properties to load
            time.sleep(4)
            # Detect new scrolling height
            new_height = dr.execute_script('return document.body.scrollHeight')
            # If new and old heights are the same that means that no new properties loaded
            if new_height == previous_height:
                # Exit the loop
                break
            # Save new height for the next loop
            previous_height = new_height

        # Lopp until the Load More button is visible
        while True:
            # Scroll to the Load More button
            try:
                scroll_to_elements = dr.find_elements(By.CLASS_NAME, 'btn-red')
                dr.execute_script("arguments[0].scrollIntoView();", scroll_to_elements[-1])
                time.sleep(2)

                load_more_button = dr.find_element(By.XPATH, "//a[text()='Load More Properties']")
            except:
                # If unable to find the button exit the loop
                break
            # Click the Load More Button
            try:
                dr.execute_script("arguments[0].click();", load_more_button)
                time.sleep(3)
            except:
                # If unable to click exit the loop
                break
        
        # Get HTML of the webpage
        page_soup = BeautifulSoup(dr.page_source, 'html.parser')
        # Get list of URLs of the individual properties
        page_listings = page_soup.find_all('a', class_='btn btn-red')

        # Iterate over every property URL
        for listing in page_listings:
            # Get text of the URL
            listing_url = self.convert_url(listing['href'])
            # Get full description of the property listing
            listing_desc = listing.parent.parent.find_previous_sibling('div', class_='panel-body')

            # Try parsing the address from the description
            try:
                listing_address = listing_desc.find('h2').getText()
            except:
                listing_address = 'Could not get address'
            # Try parsing the title from the description
            try:
                listing_title = listing_desc.find('p').getText()
            except:
                listing_title = 'Could not get title'
            # Try parsing the price from the description
            try:
                listing_price = listing_desc.find('h3').getText()
            except:
                listing_price = 'Could not get price'

            # Insert the data into the Excel file
            write_excel(self.excel_sheet, self.row_index, listing_url, listing_address, listing_title, listing_price)
            # Increase the Excel row index for the next insertion
            self.row_index += 1
    
        # Close the browser
        dr.close()

    # Format the property listing URL 
    def convert_url(self, href):
        base_url = 'https://www.gascoignehalman.co.uk'
        return base_url + href

def write_excel(sheet_name, sheet_index, url, address, title, price):
    wb = openpyxl.load_workbook(excel_file_location)
    ws = wb[sheet_name]

    ws.cell(row=sheet_index, column=1).value = url
    ws.cell(row=sheet_index, column=2).value = address
    ws.cell(row=sheet_index, column=3).value = title
    ws.cell(row=sheet_index, column=4).value = price

    wb.save(excel_file_location)
    wb.close()

def create_excel():
    global excel_file_location

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'zoopla'
    sheet2 = wb.create_sheet(title='rightmove')
    sheet3 = wb.create_sheet(title='gascoignehalman')

    sheet['A1'] = 'URL'
    sheet2['A1'] = 'URL'
    sheet3['A1'] = 'URL'

    sheet['B1'] = 'Address'
    sheet2['B1'] = 'Address'
    sheet3['B1'] = 'Address'

    sheet['C1'] = 'Title'
    sheet2['C1'] = 'Title'
    sheet3['C1'] = 'Title'

    sheet['D1'] = 'Price'
    sheet2['D1'] = 'Price'
    sheet3['D1'] = 'Price'

    excel_file_location = "Properties_{sys_time}.xlsx".format(sys_time=datetime.today().strftime('%Y-%m-%d_%H-%M'))
    wb.save(excel_file_location)
    wb.close()

def main():
    create_excel()

    zoopla_scraper = ZooplaScraper()
    zoopla_scraper.do_scrape()

    rightmove_scraper = RightMoveScraper()
    rightmove_scraper.do_scrape()

    halman_scraper = HalmanScraper()
    halman_scraper.do_scrape()

    print('Script ended')

if __name__ == "__main__":
    main()